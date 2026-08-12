from __future__ import annotations

import json

from openpyxl import load_workbook

from core import db as db_module
from ingest.adapters import fs_scan_adapter  # noqa: F401
from ingest.run_scan import run_scan
from output import static_html_adapter
from output.cli_adapter import evaluate_exit_code
from output.excel_adapter import write_workbook
from output.ndjson_export import export_ndjson
from output.report_manifest import build_manifest
from rules.engine import build_context, run_all_rules


def test_cli_exit_codes():
    assert evaluate_exit_code([], fail_on="warn") == 0
    assert evaluate_exit_code([{"severity": "info"}], fail_on="warn") == 0
    assert evaluate_exit_code([{"severity": "warn"}], fail_on="warn") == 1
    assert evaluate_exit_code([{"severity": "fail"}], fail_on="warn") == 2


def _seeded_scan(db_path, fs_scan_sample_dir):
    scan_run_id = run_scan(db_path, "fs_scan", str(fs_scan_sample_dir), since=None)
    with db_module.open_db(db_path) as conn:
        context = build_context(conn, scan_run_id, str(fs_scan_sample_dir))
        for d in run_all_rules(context):
            db_module.upsert_defect_record(
                conn,
                d.asset_id,
                d.rule_id,
                d.severity,
                d.message,
                d.suggested_fix,
                scan_run_id,
            )
    return scan_run_id


def test_cli_and_excel_share_the_same_manifest_values(
    db_path, fs_scan_sample_dir, tmp_path
):
    scan_run_id = _seeded_scan(db_path, fs_scan_sample_dir)

    with db_module.open_db(db_path) as conn:
        builder = build_manifest(conn, scan_run_id)

    manifest = builder.build()
    xlsx_path = tmp_path / "report.xlsx"
    write_workbook(builder, str(xlsx_path))

    wb = load_workbook(xlsx_path)
    ext_section = next(
        s for s in manifest["sections"] if s["section_id"] == "extension_breakdown"
    )
    ws = wb[ext_section["title"]]
    excel_values = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2)}
    manifest_values = dict(
        zip(ext_section["data"]["labels"], ext_section["data"]["values"])
    )
    assert excel_values == manifest_values


def test_ndjson_export_is_valid_bulk_format(db_path, fs_scan_sample_dir):
    scan_run_id = _seeded_scan(db_path, fs_scan_sample_dir)
    with db_module.open_db(db_path) as conn:
        lines = list(export_ndjson(conn, scan_run_id))

    assert len(lines) % 2 == 0
    for i in range(0, len(lines), 2):
        action = json.loads(lines[i])
        source = json.loads(lines[i + 1])
        assert "index" in action and "_index" in action["index"]
        assert "asset_id" in source and "check_id" in source


def test_static_html_adapter_generates_offline_report(
    db_path, fs_scan_sample_dir, tmp_path
):
    scan_run_id = _seeded_scan(db_path, fs_scan_sample_dir)
    out_path = static_html_adapter.generate(
        db_path, scan_run_id, out_dir=str(tmp_path / "reports")
    )

    html = out_path.read_text(encoding="utf-8")
    assert "http://" not in html
    assert "https://" not in html
    assert f"scan_run #{scan_run_id}" in html
